"""
generate_daily_report.py
Filter master_outage.xlsx untuk row yang "Outage Start" jatuh SEMALAM
(ikut timezone Asia/Kuala_Lumpur), export ke fail berasingan dalam
reports/daily_report_DDMMYYYY.xlsx.

Fail ni yang Power Automate flow ko akan pickup untuk hantar ke OneDrive
(bukan kerja script ni — script ni cuma HASILKAN fail report tu).

Run terus: python scripts/generate_daily_report.py
(boleh override tarikh guna --date DD/MM/YYYY untuk testing/backfill,
default = semalam)
"""

import sys
import argparse
import zoneinfo
from pathlib import Path
from datetime import datetime, timedelta

sys.path.insert(0, str(Path(__file__).resolve().parent))

from common.excel_utils import (  # noqa: E402
    get_or_create_workbook,
    write_row,
    save_workbook,
)
from common.schema import COLUMNS, FORMULA_COLUMNS  # noqa: E402

TZ = zoneinfo.ZoneInfo("Asia/Kuala_Lumpur")


def generate_daily_report(master_path: str, reports_dir: str, target_date=None):
    """
    Filter master_path untuk row "Outage Start" == target_date (default
    semalam), export ke reports_dir/daily_report_DDMMYYYY.xlsx.

    Return (report_path, row_count) — report_path None kalau master tak
    wujud, row_count 0 kalau tiada row match.
    """
    master_path = Path(master_path)
    reports_dir = Path(reports_dir)

    if target_date is None:
        target_date = (datetime.now(TZ) - timedelta(days=1)).date()

    if not master_path.exists():
        print(f"Master file tak wujud: {master_path}")
        return None, 0

    # Baca master (read-only, tak modify)
    from openpyxl import load_workbook
    wb_master = load_workbook(str(master_path), data_only=False)
    ws_master = wb_master.active

    read_columns = [c for c in COLUMNS if c not in FORMULA_COLUMNS]
    matched_rows = []

    for row_idx in range(2, ws_master.max_row + 1):
        outage_start_col = COLUMNS.index("Outage Start") + 1
        outage_start = ws_master.cell(row=row_idx, column=outage_start_col).value

        if not isinstance(outage_start, datetime):
            continue  # skip row tanpa Outage Start yang sah

        if outage_start.date() != target_date:
            continue

        row_data = {}
        for col_name in read_columns:
            col_idx = COLUMNS.index(col_name) + 1
            row_data[col_name] = ws_master.cell(row=row_idx, column=col_idx).value
        matched_rows.append(row_data)

    print(f"Tarikh disasar: {target_date} (semalam)")
    print(f"Row match: {len(matched_rows)}")

    if not matched_rows:
        return None, 0

    reports_dir.mkdir(parents=True, exist_ok=True)
    report_filename = f"daily_report_{target_date.strftime('%d%m%Y')}.xlsx"
    report_path = reports_dir / report_filename

    wb_report, ws_report = get_or_create_workbook(str(report_path))

    for i, row_data in enumerate(matched_rows, start=2):
        write_row(ws_report, i, row_data, fill_blank_only=False)

    save_workbook(wb_report, str(report_path))
    print(f"Report ditulis: {report_path}")

    return report_path, len(matched_rows)


def main():
    repo_root = Path(__file__).resolve().parent.parent

    parser = argparse.ArgumentParser(description="Generate daily outage report (Outage Start = semalam)")
    parser.add_argument(
        "--master", default=str(repo_root / "master" / "master_outage.xlsx"),
        help="Path master Excel (default: master/master_outage.xlsx)"
    )
    parser.add_argument(
        "--reports", default=str(repo_root / "reports"),
        help="Folder output report (default: reports/)"
    )
    parser.add_argument(
        "--date", default=None,
        help="Override tarikh (format DD/MM/YYYY), default = semalam. Untuk testing/backfill sahaja."
    )
    args = parser.parse_args()

    target_date = None
    if args.date:
        target_date = datetime.strptime(args.date, "%d/%m/%Y").date()

    report_path, count = generate_daily_report(args.master, args.reports, target_date)

    print()
    if report_path:
        print(f"=== Selesai: {count} row ditulis ke {report_path} ===")
    else:
        print("=== Tiada row match, report TAK dihasilkan ===")


if __name__ == "__main__":
    main()
