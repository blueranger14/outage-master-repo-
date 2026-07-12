"""
excel_utils.py
Shared function untuk baca/tulis master_outage.xlsx guna openpyxl.

Semua function di sini rujuk schema.py sebagai single source of truth
untuk column order, datetime format, dan formula.
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from common.schema import (  # noqa: E402
    COLUMNS,
    DATETIME_COLUMNS,
    FORMULA_COLUMNS,
    NUMBER_FORMATS,
    ALWAYS_OVERWRITE_FIELDS,
    BLANK_LIKE_VALUES,
)

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill


# ---------------------------------------------------------------------------
# Check sama ada value dianggap "kosong" (None/"" atau placeholder macam
# "N/A", "-", "TBC" — rujuk BLANK_LIKE_VALUES dalam schema.py)
# ---------------------------------------------------------------------------
def _is_blank_like(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip().lower() in BLANK_LIKE_VALUES:
        return True
    return False


# ---------------------------------------------------------------------------
# Column letter lookup, berdasarkan urutan dalam schema.py -> COLUMNS
# Contoh: "INC No" -> "A", "Site ID" -> "B", dst.
# ---------------------------------------------------------------------------
def _column_letter(column_name: str) -> str:
    idx = COLUMNS.index(column_name) + 1  # 1-indexed
    return get_column_letter(idx)


# ---------------------------------------------------------------------------
# Load workbook sedia ada, atau create baru dengan header row kalau tak wujud
# ---------------------------------------------------------------------------
def get_or_create_workbook(path: str):
    """
    Return (workbook, worksheet). Kalau file tak wujud, create baru dengan
    header row (bold, background kelabu) ikut COLUMNS dari schema.py.
    """
    file_path = Path(path)

    if file_path.exists():
        wb = load_workbook(file_path)
        ws = wb.active
        return wb, ws

    # Buat workbook baru
    wb = Workbook()
    ws = wb.active
    ws.title = "Master"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    ws.freeze_panes = "A2"  # header row sentiasa nampak bila scroll

    # Auto column width ringkas (anggaran ikut panjang nama column)
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = max(12, len(col_name) + 4)

    return wb, ws


# ---------------------------------------------------------------------------
# Build index {(inc_no, site_id): row_idx} SEKALI je, untuk O(1) lookup.
# PENTING untuk performance dengan data volum besar (puluhan ribu row) -
# find_existing_row() yang scan seluruh sheet setiap kali dipanggil jadi
# O(n) PER ROW = O(n^2) keseluruhan, terlalu perlahan untuk >10k rows.
# Guna build_index() sekali, then dict lookup terus (O(1)) untuk setiap
# row yang diproses.
# ---------------------------------------------------------------------------
def build_index(ws) -> dict:
    """
    Return dict {(inc_no, site_id): row_idx} untuk semua row sedia ada.
    Panggil SEKALI je di awal merge, bukan berulang kali.
    """
    inc_col = COLUMNS.index("INC No") + 1
    site_col = COLUMNS.index("Site ID") + 1

    index = {}
    for row_idx in range(2, ws.max_row + 1):
        inc_val = ws.cell(row=row_idx, column=inc_col).value
        site_val = ws.cell(row=row_idx, column=site_col).value
        if inc_val is None and site_val is None:
            continue
        index[(inc_val, site_val)] = row_idx
    return index


# ---------------------------------------------------------------------------
# Cari row number sedia ada untuk (INC No, Site ID) tertentu, None kalau baru
# ---------------------------------------------------------------------------
def find_existing_row(ws, inc_no: str, site_id: str):
    """
    Scan column INC No & Site ID (dari row 2 hingga last row), return row
    number kalau match dijumpai, None kalau tiada.

    ⚠️  PERINGATAN PERFORMANCE: function ni O(n) PER PANGGILAN (scan
    seluruh sheet). Untuk data volum besar (>1000 rows) atau proses
    banyak row dalam satu run, guna build_index() SEKALI di awal +
    dict lookup terus — JANGAN panggil find_existing_row() dalam loop
    untuk setiap row (jadi O(n^2), boleh timeout untuk data besar).
    """
    inc_col = COLUMNS.index("INC No") + 1
    site_col = COLUMNS.index("Site ID") + 1

    for row_idx in range(2, ws.max_row + 1):
        row_inc = ws.cell(row=row_idx, column=inc_col).value
        row_site = ws.cell(row=row_idx, column=site_col).value
        if row_inc == inc_no and row_site == site_id:
            return row_idx

    return None


# ---------------------------------------------------------------------------
# Tulis satu row (dict) ke row_idx tertentu dalam worksheet
# ---------------------------------------------------------------------------
def write_row(ws, row_idx: int, row_data: dict, fill_blank_only: bool = False):
    """
    Tulis row_data (dict, keys = nama column dari schema.py) ke row_idx.

    fill_blank_only=True -> untuk existing row, hanya isi cell yang KOSONG
    (tak overwrite value sedia ada). Untuk row baru, param ni diabaikan
    (semua field ditulis terus).

    Duration (Hour) ditulis sebagai FORMULA Excel, bukan value terus.
    """
    for col_name in COLUMNS:
        col_idx = COLUMNS.index(col_name) + 1
        cell = ws.cell(row=row_idx, column=col_idx)

        # --- Formula columns (Duration (Hour)) ---
        if col_name in FORMULA_COLUMNS:
            if col_name == "Duration (Hour)":
                start_letter = _column_letter("Outage Start")
                end_letter = _column_letter("Outage End")
                formula = (
                    f'=IF(AND({start_letter}{row_idx}<>"",{end_letter}{row_idx}<>""),'
                    f'{end_letter}{row_idx}-{start_letter}{row_idx},"")'
                )
                cell.value = formula
                cell.number_format = NUMBER_FORMATS.get(col_name, "General")
            continue

        # --- Field biasa ---
        if col_name not in row_data:
            continue

        new_value = row_data[col_name]

        # fill_blank_only: skip kalau cell dah ada value SEBENAR — KECUALI
        # field ni tergolong dalam ALWAYS_OVERWRITE_FIELDS, atau cell sedia
        # ada tu "blank-like" (contoh "N/A", "-", "TBC" — bukan value
        # sebenar, jadi masih boleh diganti dengan value lebih baik)
        if fill_blank_only and col_name not in ALWAYS_OVERWRITE_FIELDS:
            if not _is_blank_like(cell.value):
                continue

        if _is_blank_like(new_value):
            continue  # jangan tulis value baru yang "kosong" (None/""/N/A/dsb)

        cell.value = new_value

        if col_name in NUMBER_FORMATS:
            cell.number_format = NUMBER_FORMATS[col_name]


# ---------------------------------------------------------------------------
# Append row baru ke hujung worksheet
#
# ⚠️ PENTING: JANGAN guna ws.max_row untuk tentukan row baru bila panggil
# fungsi ni BERULANG KALI dalam loop (contoh proses ribuan row). ws.max_row
# dalam openpyxl BUKAN O(1) — setiap panggilan boleh jadi O(n) (scan
# balik worksheet), jadi loop append berulang jadi O(n^2) keseluruhan -
# lambat drastik untuk data >10k rows (contoh 48k rows boleh ambil
# BERPULUH minit, bukan saat).
#
# Untuk BULK insert (banyak row berturutan), guna next_row_idx yang
# di-track SENDIRI oleh caller (lihat merge_to_master.py / 
# merge_zip_to_master.py punya pattern), bukan panggil append_row()
# berulang kali yang each time query ws.max_row.
# ---------------------------------------------------------------------------
def append_row(ws, row_data: dict, row_idx: int = None):
    """
    Append row baru. Kalau row_idx diberi, guna terus (caller kena track
    sendiri, elak query ws.max_row berulang). Kalau tak diberi, fallback
    guna ws.max_row + 1 (selamat untuk single/occasional call, TAPI
    JANGAN guna fallback ni dalam loop besar - O(n^2)).
    """
    if row_idx is None:
        row_idx = ws.max_row + 1
    write_row(ws, row_idx, row_data, fill_blank_only=False)
    return row_idx


# ---------------------------------------------------------------------------
# Save workbook ke path
# ---------------------------------------------------------------------------
def save_workbook(wb, path: str):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ---------------------------------------------------------------------------
# Quick manual test
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import sys as _sys
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "sources"))
    from parse_email_source import parse_email_html  # noqa: E402

    with open(
        Path(__file__).resolve().parent.parent / "sources" / "test_sample_email.html",
        "r", encoding="utf-8"
    ) as f:
        html_closure = f.read()

    with open(
        Path(__file__).resolve().parent.parent / "sources" / "test_sample_alert.html",
        "r", encoding="utf-8"
    ) as f:
        html_alert = f.read()

    test_output = "/tmp/test_master_outage.xlsx"

    # 1. Alert email masuk dulu (row baru)
    wb, ws = get_or_create_workbook(test_output)
    alert_rows = parse_email_html(html_alert, subject="INC000102173805")
    for row in alert_rows:
        append_row(ws, row)
    save_workbook(wb, test_output)
    print(f"Selepas Alert: {ws.max_row - 1} rows")

    # 2. Closure email masuk (patut UPDATE row sedia ada, bukan insert baru)
    wb, ws = get_or_create_workbook(test_output)
    closure_rows = parse_email_html(html_closure, subject="INC000102173805")
    for row in closure_rows:
        existing_row_idx = find_existing_row(ws, row["INC No"], row["Site ID"])
        if existing_row_idx:
            write_row(ws, existing_row_idx, row, fill_blank_only=True)
        else:
            append_row(ws, row)
    save_workbook(wb, test_output)
    print(f"Selepas Closure: {ws.max_row - 1} rows (patut SAMA, bukan bertambah)")

    # Print hasil akhir untuk verify
    for row_idx in range(2, ws.max_row + 1):
        values = [ws.cell(row=row_idx, column=c + 1).value for c in range(len(COLUMNS))]
        print(dict(zip(COLUMNS, values)))
